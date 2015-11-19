#TODO: create sheets in mem, don't import

import unittest
import etl
from clean import clean
from openpyxl import load_workbook
from openpyxl import Workbook

#sample_wb = load_workbook('/Users/ewanog/code/git_repos/nepal-earthquake/shelter/etl-extravaganza/clean_test.xlsx', data_only = True)
#train = sample_wb.get_sheet_by_name('Training')
#ref = sample_wb.get_sheet_by_name('Reference')
#date_wb = load_workbook('/Users/ewanog/code/git_repos/nepal-earthquake/shelter/etl-extravaganza/date_test.xlsx', data_only = True)
#dist = date_wb.get_sheet_by_name('dates')

class TestClean(unittest.TestCase):

    def test_algo18_fail(self):
        db = Workbook().active
        ref = Workbook().active

        db.append(('DD - Start', 'MM - Start', 'YYYY - Start', '', 'psych!!'))
        db.append(('12', '12', '1990', 'r belong!!'))
        db.append(('12', '12', '2120', 'r belong!!'))

        self.assertEqual(clean.algo18(db,ref)[2][1], ["1990-12-12 00:00:00 @ row 2"])

    def test_algo18_pass(self):
        db = Workbook().active
        ref = Workbook().active

        db.append(('DD - Start', 'MM - Start', 'YYYY - Start', '', 'psych!!'))
        db.append(('12', '12', '2020', 'r belong!!'))

        self.assertEqual(clean.algo18(db,ref)[2][1], [])

    def test_algo19_fail(self):
        db = Workbook().active
        ref = Workbook().active

        db.append(('etc','DD - Start', 'MM - Start', 'YYYY - Start', 'DD - End', 'MM - End', 'YYYY - End', 'etc'))
        db.append(('**','12', '12', '1990', '12', '12', '1990'))
        db.append(('**','12', '12', '2990', '12', '12', '1990'))

        self.assertEqual(clean.algo19(db,ref)[2][1], ["2990-12-12 00:00:00 @ row 3"])

    def test_algo19_pass(self):
        db = Workbook().active
        ref = Workbook().active

        db.append(('etc','DD - Start', 'MM - Start', 'YYYY - Start', 'DD - End', 'MM - End', 'YYYY - End', 'etc'))
        db.append(('**','12', '12', '1980', '12', '12', '1990'))
        db.append(('**','12', '12', '1890', '12', '12', '1990'))

        self.assertEqual(clean.algo19(db,ref)[2][1], [])

    def test_algo20_fail(self):
        db = Workbook().active
        ref = Workbook().active

        db.append(('etc','DD - End', 'MM - End', 'YYYY - End', 'Activity Status'))
        db.append(('**','12', '12', '1890', 'breakit'))
        db.append(('**','12', '12', '1890', 'completed'))

        self.assertEqual(clean.algo20(db,ref)[2][1], [("1890-12-12 00:00:00 row 2")])

    def test_algo20_pass(self):
        db = Workbook().active
        ref = Workbook().active

        db.append(('etc','DD - End', 'MM - End', 'YYYY - End', 'Activity Status'))
        db.append(('**','12', '12', '1890', 'it is completed'))
        db.append(('**','12', '12', '1890', 'completed'))

        self.assertEqual(clean.algo20(db,ref)[2][1], [])

if __name__ == '__main__':
    unittest.main()


    
