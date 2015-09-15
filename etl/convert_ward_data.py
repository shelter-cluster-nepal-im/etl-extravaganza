from openpyxl import Workbook
import etl
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

def get_vals(ws):
    w = etl.pull_wb("/Users/ewanog/Downloads/ward.xlsx", "local")
    cur_y = 47
    found = True
    vals = []
    #dist = 'A'
    #vdc = 'F'
    ward = 'A'
    pop = 'D'
    hh = 'B'
    dist = ws['A44'].value.split()[2]
    #
    while found:
        vdc = ws['F'+str(cur_y-3)].value.split()[2]
        if ws['F'+str(cur_y-3)].value.split()[3][0]!='[':
            vdc+= ' ' + ws['F'+str(cur_y-3)].value.split()[3]

        i = 0
        while isinstance(ws['A'+str(cur_y)].value,int):
            vals.append([dist, vdc, ws[ward+str(cur_y)].value,ws[pop+str(cur_y)].value,ws[hh+str(cur_y)].value])
            cur_y+=1

        cur_y+=7
        if ws['A'+str(cur_y)].value != 1 and vdc != 'Bhusafeda':
            found = False
        else:
            vdc = ws['F'+str(cur_y-3)].value.split()[2]

    return vals

def pr(r):
    for v in r:
        for i in v:
            print str(i) + ",",
        print

def match_vdcs():
    #notes...
    #1651 VDCs without exact match


    w = etl.pull_wb("/Users/ewanog/Downloads/ward_level_data.xlsx", "local")
    #w = etl.pull_wb("/Users/ewanog/Downloads/test.xlsx", "local")
    #from cbs: col 1 (dist), col 2 (vdc), col 3 (ward)
    #from new: H (dist), N (VDC), O (ward)
    cbs = w.get_sheet_by_name("cbs")
    new = w.get_sheet_by_name("new")
    cbs_dist = etl.get_values(cbs.columns[0][1:])
    cbs_vdc = etl.get_values(cbs.columns[1][1:])
    cbs_ward = etl.get_values(cbs.columns[2][1:])

    new_dist = etl.get_values(new.columns[7][1:])
    new_vdc = etl.get_values(new.columns[13][1:])
    new_ward = etl.get_values(new.columns[14][1:])
    ns = set(new_vdc)
    #see vdc matches

    cb_zip = zip(cbs_dist, cbs_vdc)
    new_zip = zip(new_dist, new_vdc)


    print len(set(etl.colvals_notincol(cbs,'B',new,'N')))
    print len(set(cbs_vdc))
    for v in set(etl.colvals_notincol(cbs,'B',new,'N')):
        r = process.extract(v, [r for r in new_zip if r[1] == ])
        print r[0][1]


if __name__ == '__main__':
    read = False
    if read:
        for ws in w.worksheets:
            if ws['A44'].value:
                r = get_vals(ws)
                print r[-1][0]
                #pr(r)
    else:
        match_vdcs()



