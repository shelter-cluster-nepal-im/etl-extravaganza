import unittest
import etl
import metadata
from openpyxl import Workbook

class TestMetadata(unittest.TestCase):

    #meta needs: ref_sht, acc_sht, codes_sht
    #["Priority", "Hard to Reach Access Methods", ]

    def test_priority(self):
        db = Workbook().active
        db.append(('Empty','District', 'Priority'))
        db.append(('Empty','ispri'))
        db.append(('Empty','isnotpri'))

        ref = Workbook().active
        ref.append(('Other','Other','Priority Districts'))
        ref.append(('etc','etc','ispri'))

        m = metadata.Meta(ref, None, None)
        r = m.get('Priority', db)
        self.assertEquals(r, ['TRUE','FALSE'])

    def test_reach(self):
        db = Workbook().active
        db.append(('District', 'VDC / Municipalities'))
        db.append(('d1','vdc1'))
        db.append(('xx','yy'))

        acc = Workbook().active
        acc.append(('DistrictVDC Concatenation', 'NeKSAP ACCESS'))
        acc.append(('d1vdc1', 'amethod'))

        m = metadata.Meta(None, acc, None)
        r = m.get('Hard to Reach Access Methods', db)
        self.assertEquals(r, ['amethod',''])

    def test_hub(self):
        db = Workbook().active
        db.append(('blank', 'District'))
        db.append(('e','d1'))
        db.append(('i','d2'))

        ref = Workbook().active
        ref.append(('Priority Districts', 'Shelter Cluster Hubs'))
        ref.append(('d4', 'notin'))
        ref.append(('d2', 'madpri'))

        m = metadata.Meta(ref, None, None)
        r = m.get('Shelter Cluster Hub', db)
        self.assertEquals(r, ['','madpri'])


    def test_update(self):
        db = Workbook().active
        db.append(('Last Update',''))
        db.append(('date',''))
        db.append(('moredates',''))

        m = metadata.Meta(None, None, None)
        r = m.get('Last Update', db)
        self.assertEquals(r, ['date','moredates'])


    def test_dist_hlcit(self):
        db = Workbook().active
        db.append(('District'))
        db.append(('thisd',''))
        db.append(('nope',''))
        db.append(('kendrik',''))

        ref = Workbook().active
        ref.append(('Admin1_District', 'Admin1_P-Code'))
        ref.append(('kendrik', 'lamar'))
        ref.append(('thisd', 'thatd'))

        m = metadata.Meta(ref, None, None)
        r = m.get('District HLCIT Code', db)
        self.assertEquals(r, ['thatd','','lamar'])


    def test_vdc_hlcit(self):
        db = Workbook().active
        db.append(('District', 'VDC / Municipalities'))
        db.append(('d1','v1'))
        db.append(('e','i'))

        ref = Workbook().active
        ref.append(('Admin1 + Admin2 Concatenation', 'Admin2_HLCIT_CODE'))
        ref.append(('d1v1', 'code'))

        m = metadata.Meta(ref, None, None)
        r = m.get('VDC / Municipality HLCIT Code', db)
        self.assertEquals(r, ['code',''])

    def test_act_cat(self):
        db = Workbook().active
        db.append(('Action description',''))
        db.append(('in',''))
        db.append(('notin',''))

        ref = Workbook().active
        ref.append(('SC Categories','UN Categories'))
        ref.append(('in', 'res'))

        m = metadata.Meta(ref, None, None)
        r = m.get('UNOCHA Activity Categories', db)
        self.assertEquals(r, ['res',''])

    def test_uid(self):
        db = Workbook().active
        db.append(("Implementing agency", "Local partner agency" , "District",
                "VDC / Municipalities", "Municipal Ward", "Action type",
                "Action description", "# Items / # Man-hours / NPR",
                "Total Number Households"))
        db.append(('1','2','3','4','5','6','7','8','9'))

        m = metadata.Meta(None, None, None)
        r = m.get('UID', db)
        self.assertEquals(r, ['123456789'])

if __name__ == '__main__':
    unittest.main()
