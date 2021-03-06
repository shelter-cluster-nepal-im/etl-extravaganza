"""Module for extracting data from dropbox and aggregating"""
#TODO: better way for specifyign which folder to pull from (latest?)
##pull from rules based text file
##log put on dbox, make better
##test individual cleaning
##mvoe current logs to an 'old' file 
##put in run params: exclude
##change to pull down all WSs at same time as opposed to iterating  
##get rid of spcifycing column for uid, should just be one after header
##logs output to local or dbox
##single method to import all file(s) from a path

from collections import Counter
from copy import copy
import csv
import datetime
import dropbox
import clean
import metadata
import os
import cStringIO
import re
from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string
from openpyxl import Workbook
import openpyxl.writer.excel as wrtex
import time
import click
from os import listdir
from os.path import isfile, join
from openpyxl.utils import get_column_letter
from itertools import groupby
from openpyxl.styles import PatternFill
from io import BytesIO

#dropbox setup
db_access = os.environ['db_access']
client = dropbox.client.DropboxClient(db_access)

@click.command()
@click.option('--act', help='what action will we perform', type = click.Choice(['cons','clean','split']))
@click.option('--src', help='local files or on dropbox?', type = click.Choice(['db','local']))
@click.option('--path', help='file path (pull all xls or xlsx files in it with a flag (-C currently)')
@click.option('--db', help='db file if consolidating or splitting')
@click.option('--test', help='are we testing?', is_flag = True)

def iterate_reports(act, src, path, db, test):
    """cycle through all reports contained in dbox directory"""

    if act in ('clean', 'cons'):
        file_list = get_file_list(path, src)
        file_list = clean_exclude(file_list)

        if test == True:
            file_list = [file_list[0]]

        for v in file_list:
            print v

        #pull down all workbooks
        wbs = []
        for f in file_list:
            #pull down workbook from specified directory
            print "Pulling: " + f
            wb_current = pull_wb(f, src, True)

            #check to see if properly formatted
            if wb_format(wb_current):
                print 'Appending: ' + f
                wbs.append((wb_current,f))

            else:
                #put in malformatted folder
                print 'Malformatted ' + f
                send_path = path + '/old_format_or_incorrect/' + f.rsplit('/', 1)[1]
                send_wb(send_path, wb_current, src)

    #clean, consolidate or split
    if act == 'clean':
        #clean workboooks
        for wb in wbs:
            print "Cleaning: " + wb[1]
            clean_file(wb[0], wb[1], src)

        emit_log(path, src)

    elif act == 'cons':
        #consolidate
        to_send = consolidate(pull_wb(db, src, True), wbs, src)

        send_wb(path + '/consolidated.xlsx', to_send, src)

    elif act == 'split':
        #consolidate - path is base dir where folder for splits will be made
        dbsht = get_unsure_ws(pull_wb(db, src, True),['Database','Distributions'])

        split(dbsht, path, src)

def get_unsure_ws(wb, names):
    """pull out ws from list in order for checking for multiple names"""
    for n in names:
        if n in wb.sheetnames:
            return wb.get_sheet_by_name(n)

def clean_exclude(file_list):
    new_list = []
    for v in file_list:

        if v.split('/')[-1].startswith('C-') or v.split('/')[-1].startswith('C -'):
            new_list.append(v)
        else:
            print 'Excluding: ' + v

    return new_list


def find_none_ws_count(ws):
    cnt = 0
    for r in ws.rows:
        if not none_row(r):
            cnt+=1
        else:
            break

    return cnt

def get_all_values_from_ws(ws, r_or_c):
    """return an array of arrays f.e column in a ws"""

    if r_or_c == 'r':
        return [get_values(r) for r in ws.rows if r[0].value is not None]

    if r_or_c == 'c':
        return [get_values(c) for c in ws.columns if c[0].value is not None]


def split_get_sheets(db):
    """return a tuple with agname and array of arrays of sheet contents"""
    ddict = get_all_values_from_ws(db)
    ag_loc = column_index_from_string(find_in_header(db, "Implementing Agency"))-1
    ddict = sorted(ddict[1:], key=lambda k: k[ag_loc])

    #group by and get individual sheets
    to_ret = []
    for key, group in groupby(ddict, lambda k: k[ag_loc]):
        cur_ag = []
        for r in group:
            cur_ag.append(r)

        to_ret.append((key,cur_ag))

    return to_ret

def get_template(src, strip):
    if src == 'db':
        return pull_wb('/2015 Nepal EQ/04 im/06 team folders/EO_Folder/EO_scripts/reportingtemplate_sheltercluster.xlsx', 'db', strip)
    elif src == 'local':
        return pull_wb('/Users/ewanog/Dropbox (GSC)/2015 Nepal EQ/04 IM/Reporting/Database_&_Template/EO_scripts/reportingtemplate_sheltercluster.xlsx','local', strip)

def split(db, path, src):
    """split a db file into agency specific xlsx"""
    ws_list = split_get_sheets(db)
    template = get_template(src, True)

    #trim down to just essential columns
    iav = column_index_from_string(find_in_header(db,'Implementing Agency'))-1
    acv = column_index_from_string(find_in_header(db,'Additional comments'))

    #remove additional sheets in template
    for s in template.sheetnames:
        if s != 'Distributions':
            template.remove_sheet(template.get_sheet_by_name(s))

    header_vals = get_values(template.get_sheet_by_name('Distributions').rows[0])
    for ws in ws_list:
        send = template
        send_sheet = send.get_sheet_by_name('Distributions')

        if len(ws[1][iav:acv]) != len(send_sheet.rows[0]):
            Exception('Template header doesnt match!')

        for i,r in enumerate(ws[1][1:]):
            for i_i, v in enumerate(r):
                send_sheet.cell(row =i+2, column = i_i + 1).value = v

        send_wb(path+'split/'+ ws[0] + ' - ' + datetime.datetime.now().strftime('%d%m%y') + '.xlsx',send, src)

        #workaround for ws copying behavior - remove all values
        for r in send_sheet.rows[1:]:
            for v in r:
                v.value = None



        send_sheet = None


def print_sheet(sht):
    """print out all values in a sheet"""
    for r in sht.rows:
        print get_values(r)
        print

def copy_sheet_vals(to, fro):
    """read in two sheets and copy values from the second one"""
    for c_ind, c_v in enumerate(fro.columns):
        c_ind+=1
        for r_ind, r_v in enumerate(c_v):
            #adjust for openpy locs
            r_ind+=1
            to.cell(row = r_ind, column = c_ind).value = fro.cell(row = r_ind, column = c_ind).value

    return to

def copy_sheet_styles(to, fro):
    """read in two sheets and copy styles from the second one"""
    for c_ind, c_v in enumerate(fro.columns):
        c_ind+=1
        for r_ind, r_v in enumerate(c_v):
            #adjust for openpy locs
            r_ind+=1
            to.cell(row = r_ind, column = c_ind).style = fro.cell(row = r_ind, column = c_ind).style

    return to


def add_vals(sht, v_add, start, r_or_c):
    cc = sht.cell(start)

    if r_or_c == 'r':
        c_add = 1
        r_add = 0

    if r_or_c == 'c':
        r_add = 1
        c_add = 0

    for i,v in enumerate(v_add):
        cc.value = v
        #prevent creation of blank cell
        if i != len(v_add)-1:
            cc = cc.offset(row = r_add, column = c_add)

    return sht


def add_metadata(sht, src):
    """add in metadata columns to the end of a sheet"""
    cols_to_add = ["Priority", "Hard to Reach Access Methods", "Shelter Cluster Hub", \
                   "District HLCIT Code", "VDC / Municipality HLCIT Code", "UID"]

    #Trainings does not have this column
    if sht.title == 'Distributions':
        cols_to_add.append("UNOCHA Activity Categories")


    #find the cell value where we should start inserting our metadata
    cur_start = sht.cell(find_last_value(sht, 'A', 'r')).offset(row = 1, column = 1)

    #now, add in metadata column names
    sht = add_vals(sht, cols_to_add, sht.cell(find_last_value(sht, 'A', 'r')).coordinate, 'r')

    #pull down relevant sheets and create metadata object
    template = get_template(src, True)
    ref_sht = template.get_sheet_by_name('Reference')
    acc_sht = template.get_sheet_by_name('Access_HumPovIndex')
    codes_sht = template.get_sheet_by_name('VDC_MapCodes')
    m = metadata.Meta(ref_sht, acc_sht, codes_sht)

    for c in cols_to_add:
        sht = add_vals(sht, m.get(c, sht), cur_start.coordinate, 'c')
        cur_start = cur_start.offset(column = 1)

    return sht


def add_meta_colors(ws, template):
    """add colors to metadata columns (we assume everything beyond 'Additional Comments'"""
    cur_cell = ws.cell(find_in_header(ws, 'Additional Comments') + '1').offset(column = 1)

    while cur_cell.value is not None:
        cur_cell.fill = PatternFill(fgColor = 'F09132', patternType = 'solid')
        cur_cell = cur_cell.offset(column = 1)

    return ws


def consolidate(baseline_ret, wsl, src):
    """consolidate Distributions and Trainings"""
    ret = Workbook()
    ret.remove_sheet(ret.get_sheet_by_name('Sheet'))
    ret.create_sheet(1,'Distributions')
    ret.create_sheet(2,'Trainings')

    template = get_template(src, True)
    dist_template = template.get_sheet_by_name('Distributions')
    train_template = template .get_sheet_by_name('Trainings')

    print 'Getting metadata for Dist'
    dist = ret.get_sheet_by_name('Distributions')
    dist = copy_sheet_vals(dist,consolidate_specfic(baseline_ret, wsl, 'Distributions').get_sheet_by_name('Distributions'))
#    dist = copy_sheet_styles(dist, dist_template)
    dist = add_metadata(dist, src)
#    dist = add_meta_colors(dist, dist_template)

    print 'Getting metadata for Training'
    train = ret.get_sheet_by_name('Trainings')
    train = copy_sheet_vals(train, consolidate_specfic(baseline_ret, wsl, 'Trainings').get_sheet_by_name('Trainings'))
#    train = copy_sheet_styles(train, train_template)
    train = add_metadata(train, src)
#    train = add_meta_colors(train, train_template)

    return ret

def clean_output(cons):
    """convert appropriate columns to numeric values and format dates properly"""
    num_cols = ["# Items / # Man-hours / NPR", "Total Number Households", "Average cost per households (NPR)",\
                    "Female headed households", "Vulnerable Caste / Ethnicity households", 'Duration of each session (hours)',  \
                'Amount Paid to Participants (NPR per participant)',    'Total Cost Per Training',\
                'Total Participants (Individuals)', 'Males',    'Females',  'Third Gender', 'Elderly (60+)',    \
                'Children (under 18)',  'Persons with Disabilities',    'Vulnerable Caste or Ethnicity',    \
                'Female Headed Households (if applicable)', \
                'DD - Start', 'MM - Start', 'YYYY - Start', 'DD - End', 'MM - End', 'YYYY - End' ]


    for col in cons.columns:
        if col[0].value in num_cols:
            for v in col:
                try:
                    v.value = int(v.value)
                except:
                    pass

    return cons

def convert_date(from_dt, from_dt_fmt, to_dt_fmt):
    """take in a date and convert it to approriate format"""
    return datetime.datetime.strptime(from_dt, from_dt_fmt).strftime(to_dt_fmt)

def consolidate_specfic(baseline_ret, wsl, which_sheet):
    """consolidate a given sheet type and remove old data"""

    cons_wb = Workbook()
    cons = cons_wb.active
    cons.title = which_sheet

    to_add = []
    ag_skip = []
    cnts = {}

    baseline_ret = baseline_ret.get_sheet_by_name(which_sheet)

    #trim baseline to be between Implementing agency and additional comments
    iav = column_index_from_string(find_in_header(baseline_ret,'Implementing Agency'))-1
    acv = column_index_from_string(find_in_header(baseline_ret,'Additional comments'))
    luv = column_index_from_string(find_in_header(baseline_ret,'Last Update'))
    baseline = Workbook().active

    for i,r in enumerate(baseline_ret.rows):
        #insert header
        if i == 0:
            baseline.append(get_values(r[iav:acv]) + get_values([r[luv-1]]))
        #else, add last update dates. try and format if not null
        else:
            if r[luv-1].value is not None:
                try:
                    baseline.append(get_values(r[iav:acv]) + [r[luv-1].value.strftime('%d/%m/%Y')])
                except Exception, e:
                    baseline.append(get_values(r[iav:acv]) + [convert_date(r[luv-1].value, '%d/%m/%Y', '%d/%m/%Y')])

    db_ialoc = find_in_header(baseline,'Implementing Agency')
    base_count = Counter(get_values(baseline.columns[column_index_from_string(db_ialoc)-1]))
    cons.append(get_values(baseline_ret.rows[0][iav:acv]) + ['Last Update'])


    #iterate through each sheet and find entries to be added
    for ws in wsl:
        cd = Workbook().active

        #make sure col names are correct
        if 'Distribution' in ws[0].sheetnames:
            ws[0].get_sheet_by_name('Distribution').title = 'Distributions'
        if 'Training' in ws[0].sheetnames:
            ws[0].get_sheet_by_name('Training').title = 'Trainings'

        #trim down to just essential columns
        print ws[0].worksheets
        iav = column_index_from_string(find_in_header(ws[0].get_sheet_by_name(which_sheet),'Implementing Agency'))-1
        acv = column_index_from_string(find_in_header(ws[0].get_sheet_by_name(which_sheet),'Additional comments'))
        for r in ws[0].get_sheet_by_name(which_sheet).rows:
            cd.append(get_values(r[iav:acv]))

        wsialoc = find_in_header(cd,'Implementing Agency')
        ag_name = cd[str(wsialoc + '2')].value
        if ag_name is None:
            print '***ERROR: Agency name missing for ' + ws[1]

        #check if headers match
        wslen = len(cd.rows[0])+1
        hdlen = len(cons.rows[0])
        if wslen != hdlen:
            print '***ERROR: Non-matching header for: ' + xstr(cd[str(wsialoc + '2')].value)
            print get_values(cd.rows[0])
            print get_values(cons.rows[0])

        else:
            #check to see if agency is in list and if it is > 80 pct
            if base_count.has_key(ag_name) and ag_name.lower() != 'government':
                if find_none_ws_count(cd) < base_count[ag_name]*.8:
                    print '***WARNING: ' + ag_name + ' is less than 80 pct'
                else:
                    if ag_name.lower() != 'government':
                        ag_skip.append(xstr(ag_name))

                    cnts[ag_name] = [str(find_none_ws_count(cd)-1),0]
                    for r in cd.rows[1:]:
                        if none_row(r):
                            break
                        else:
                            to_add.append(get_values(r))
            else:
                cnts[ag_name] = [str(find_none_ws_count(cd)-1),0]
                for r in cd.rows[1:]:
                    if none_row(r):
                        break
                    else:
                        to_add.append(get_values(r))

    #create master file
    #this could be better with grouping
    cs = ''
    for v in baseline.rows[1:]:
        if xstr(v[column_index_from_string(db_ialoc)-1].value) not in ag_skip:
            if none_row(v):
                break
            else:
                cons.append(v)

        else:
            if cs != v[column_index_from_string(db_ialoc)-1].value:
                cnts[xstr(v[column_index_from_string(db_ialoc)-1].value)][1] = xstr(base_count[v[column_index_from_string(db_ialoc)-1].value])
                cs = xstr(v[column_index_from_string(db_ialoc)-1].value)

    #add in new agency info
    for v in to_add:
        cons.append(v + [datetime.datetime.now().strftime('%d/%m/%Y')])



    #convert columns to numbers
    cons = clean_output(cons)

    #print_sheet(cons)

    #print info
    print 'Insert/Delete Info...'
    print 'Agency : Insert Count : Delete Count'

    for k,v in cnts.iteritems():
        if k is not None:
            print k + ' : ' + str(v[0]) + ' : ' + str(v[1])

    return cons_wb


def none_row(val):
    merge =''
    for v in get_values(val):
        merge+=v
    if len(merge) == 0:
        return True

def get_values(r, **kargs):
    """returns values of a row or a column - note dates are specifically formatted"""
    #TODO: parameterize date format?
    ret = []
    for v in r:
        if xstr(v.value) == 'None':
            if kargs.has_key('setnull') and kargs['setnull']:
                ret.append(None)
            else:
                ret.append('')
        elif isinstance(v.value, datetime.datetime):
            ret.append(v.value.strftime('%d/%m/%Y'))
        else:
            ret.append(xstr(v.value))

    return ret

def send_wb(path, wb, src):
    print 'Sending... ' + path
    if src == 'db':
        client.put_file(path, wrtex.save_virtual_workbook(wb))

    elif src == 'local':
        if not os.path.exists(path.rsplit('/', 1)[0]):
            os.makedirs(path.rsplit('/', 1)[0])
        wb.save(path)

    print 'Sent!'



def wb_format(wb):
    """check to see if a report is correct and in the new report format"""
    must_contain = ['Distributions', 'Training', 'Reference']

    match_count = 0
    for s in wb.worksheets:
        if s.title in must_contain:
            match_count+=1

    if match_count < 3:
        return False
    else:
        return True

def clean_file(wb, path, src):
    """cycle through a report and apply cleaning algorithms"""

    #get our two sheets
    db = wb.get_sheet_by_name('Distributions')
    ref = wb.get_sheet_by_name('Reference')

    #setup log
    rname =  path.rsplit('/', 1)[1]


    #####do edit stuff
    #algos return db, ref, message

    #algo1
    db, ref, message = clean.clean.algo1(db,ref)
    report_a_log(message, rname)

    #algo2
    db, ref, message = clean.clean.algo2(db,ref)
    report_a_log(message, rname)

    #algo3
#    db, ref, message = clean.clean.algo3(db,ref)
#    report_a_log(message, rname)

    #algo4
    db, ref, message = clean.clean.algo4(db,ref)
    report_a_log(message, rname)

    #algo5
    db, ref, message = clean.clean.algo5(db,ref)
    report_a_log(message, rname)

    #algo6
    db, ref, message = clean.clean.algo6(db,ref)
    report_a_log(message, rname)

    #algo7
    db, ref, message = clean.clean.algo7(db,ref)
    report_a_log(message, rname)

    #algo8
    db, ref, message = clean.clean.algo8(db,ref)
    report_a_log(message, rname)

    #algo9
    db, ref, message = clean.clean.algo9(db,ref)
    report_a_log(message, rname)

    #algo10
    db, ref, message = clean.clean.algo10(db,ref)
    report_a_log(message, rname)

    #algo11
    db, ref, message = clean.clean.algo11(db,ref)
    report_a_log(message, rname)

    #algo12
    db, ref, message = clean.clean.algo12(db,ref)
    report_a_log(message, rname)

    #algo13
    db, ref, message = clean.clean.algo13(db,ref)
    report_a_log(message, rname)

    #algo14
    db, ref, message = clean.clean.algo14(db,ref)
    report_a_log(message, rname)

    #algo15
    db, ref, message = clean.clean.algo15(db,ref)
    report_a_log(message, rname)

    #algo16
    db, ref, message = clean.clean.algo16(db,ref)
    report_a_log(message, rname)

    #algo17
    db, ref, message = clean.clean.algo17(db,ref)
    report_a_log(message, rname)

    #algo18
    db, ref, message = clean.clean.algo18(db,ref)
    report_a_log(message, rname)

    #algo19
    db, ref, message = clean.clean.algo19(db,ref)
    report_a_log(message, rname)

    #algo20
    db, ref, message = clean.clean.algo20(db,ref)
    report_a_log(message, rname)

#    send_wb(path.rsplit('/', 1)[0] + '/edited/' + path.rsplit('/', 1)[1], wb, src)
#    print 'uploaded! ' + path.rsplit('/', 1)[0] + '/edited/' + path.rsplit('/', 1)[1]

def save_file(file, path, src):
    """write a file to either db or local"""
    if src == 'db':
        client.put_file(path + "/clean_out/log_" + time.strftime("%m-%d-%y_%H-%M-%S") +'.csv', file)

    elif src == 'local':
        pass
        #must implement

current_log = []
def emit_log(path, src):
    global current_log
    t = BytesIO()

    csv_out = csv.writer(t, delimiter=',')
    csv_out.writerow(['File Name', 'Issue', 'Instances'])
    for log in current_log:
        #if we have contents to log
        if len(log[1][1]) > 0:
	    #must check to see if char length exceeds excel max cell contents
	    if sum([len(x) for x in log[1][1]]) > 32700:
		log[1][0] += ' Warning: Too many errors. Please fix the ones shown and re-run!'	        
		
		c = 0
		for i,v in enumerate(log[1][1]):
		    if c + len(v) > 32700:
			break
		    else:
			c += len(v) + 3
		log[1][1] = log[1][1][0:i]
	   
	    csv_out.writerow([log[0]] + [log[1][0]] + [" | ".join(xstr(x) for x in log[1][1])])

    save_file(t, path, src)
    t.close()
    current_log = []

def report_a_log(log_value, path):
    """write out contents for a given log - creates new entry if a new path is given"""
    #todo: this is gross
    global current_log

    current_log.append([path, log_value])



def find_in_header(sheet, find_val):
    """find the coordinate of a value in header (assumes header is in row 1)"""
    #TODO:if we don't find anything, try fuzzy match
    for row in sheet.iter_rows('A1:' + find_last_value(sheet,'A','r')):
        for cell in row:
            if cell.value.replace(' ','').replace('\n','').lower() == find_val.replace(' ','').replace('\n','').lower():
                return cell.column

    #check if it's a date with (Actual or Planned)
    if 'date' in find_val.lower():
        find_val+='(Actual or Planned)'
        for row in sheet.iter_rows('A1:' + find_last_value(sheet,'A','r')):
            for cell in row:
                if cell.value.replace(' ','').replace('\n','').lower() == find_val.replace(' ','').replace('\n','').lower():
                    return cell.column


    print 'No header found for: ' + find_val
    return None

def colvals_notincol(sheet_val,col_val,sheet_ref,col_ref):
    """return values from a column that are NOT in a reference column"""
    not_in = []
    to_search = []

    #create an array from sheet_ref with values to be searched (as opposed to nested loops)
    #iter_rows syntax: sheet.iter_rows('A1:A2')
    for row in sheet_ref.iter_rows(col_ref + "2:" +
        find_last_value(sheet_ref, col_ref, 'c')):

        for cell in row:
                try:
                    to_search.append(xstr(cell.value))
                except:
                    to_search.append(str(cell.value))

    #now search through vals and see if they're present
    for row in sheet_val.iter_rows(col_val + "2:" +
        find_last_value(sheet_val, col_val, 'c')):

        for cell in row:
            if xstr(cell.value) not in to_search:
                try:
                    not_in.append(xstr(cell.value))
                except:
                    not_in.append(str(cell.value))

    return not_in



def find_last_value(sheet, start_location, r_or_c):
    """find position of last value in a given row or column - if row, we assume header"""
    #extract form r_or_c if we should be iterating a row or column

    if r_or_c == 'c':
        loc = 0
        for v in get_values(sheet.columns[column_index_from_string(start_location)-1]):
            if v == '':
                break
            else:
                loc+=1
        return start_location + str(loc)

    elif r_or_c == 'r':
        return get_column_letter(sheet.max_column) + '1'
    else:
        raise Exception("r_or_c must be r or c!")


def pull_wb(location, src, strip):
    """return an excel file from either local or source"""
    if strip:
        w = wb_strip(location, src)
    else:
        if src == 'db':
            w = load_workbook(pull_from_db(location))
        else:
            w = load_workbook(location)

    return w

def wb_strip(location, src):
    if src == 'db':
        w = load_workbook(pull_from_db(location), read_only = True, data_only = True)
    else:
        w = load_workbook(location, read_only = True, data_only = True)

    new_wb = Workbook()
    if len(new_wb.worksheets) > 0:
        new_wb.remove_sheet(new_wb.worksheets[0])

    for v in w.worksheets:
        print 'list'
        print w.worksheets
        cur_w = new_wb.create_sheet(1, v.title)
        for r in v.rows:
            for v in r:
                if v.value is not None:
                    cur_w[v.coordinate] = v.value

    print "Pulled: " + location
    print "With tabs: " + str(new_wb.get_sheet_names())

    return new_wb

def pull_from_db(path):
    """pull a file from dropbox"""
    to_ret = cStringIO.StringIO()

    with client.get_file(path) as f:
        to_ret.write(f.read())
    f.close()

    return to_ret

def get_file_list(path, src):
    """return file list from local or db"""
    if src == 'db':
        meta = client.metadata(path, list=True)
        return [str(f['path']) for f in meta['contents'] if re.search('xls|xlsx$',str(f))]

    elif src == 'local':
        return [str(path +'/' + f) for f in listdir(path) if isfile(join(path,f)) and re.search('xls|xlsx$',str(f))]

def fuzzy_match_col(col, vals):
    pass

def xstr(conv, **kargs):
    """return a more battle teststed encoded string"""
    if (conv == 'None' or conv is None or conv == '') and kargs.has_key('setnull') and kargs['setnull']:
        return None

    try:
        return str(conv.encode('utf8'))
    except:
        return str(conv)

if __name__ == '__main__':
    iterate_reports()

