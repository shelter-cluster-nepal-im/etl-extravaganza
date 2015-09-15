import unittest
import etl
import xls_to_sql
from sqlalchemy import Column, Integer, String, Float
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import column_index_from_string
import os
Base = declarative_base()
engine = create_engine(os.environ['dbk'])
Base.metadata.bind = engine
DBSession = sessionmaker(bind=engine)
session = DBSession()
Base.metadata.create_all(engine)


class TestDBb(unittest.TestCase):

    def test_get_sheets(self):
        wb = Workbook()
        wb.remove_sheet(wb.worksheets[0])
        wb.create_sheet(1,'sheet1')
        wb.create_sheet(2,'sheet2')
        wb.create_sheet(3,'sheet3')
        ret = xls_to_sql.get_sheets([wb], ('sheet1', 'sheet3'))
        self.assertEqual([v.title for v in xls_to_sql.get_sheets([wb], ('sheet1', 'sheet3'))] \
                         , ['sheet1','sheet3'])


class test(Base):
    __tablename__ = 'test'
    targeting = Column(String(250), primary_key=True)
    quantity = Column(Integer)
    total_hh = Column(Integer)
    float = Column(Float)

if __name__ == '__main__':
    unittest.main()
