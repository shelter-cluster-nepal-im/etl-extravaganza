import unittest
import etl
from openpyxl import load_workbook
from openpyxl import Workbook
from collections import Counter
import os
from openpyxl.styles import PatternFill


sample_wb_new_format = load_workbook('/Users/ewanog/code/nepal-earthquake/shelter/etl-extravaganza/clean_test.xlsx', data_only = True)
sample_wb = load_workbook('/Users/ewanog/code/nepal-earthquake/shelter/etl-extravaganza/etl_test.xlsx')
db = sample_wb.get_sheet_by_name('Database')
ref = sample_wb.get_sheet_by_name('Reference')

class TestEtl(unittest.TestCase):

    def test_wb_format_false(self):
        self.assertEqual(etl.wb_format(sample_wb), False)

    def test_wb_format_true(self):
        self.assertEqual(etl.wb_format(sample_wb_new_format), True)

    def test_find_last_value_row(self):
        self.assertEqual(etl.find_last_value(ref, 'A', 'r'), 'Y1')

    def test_find_last_value_column(self):
        self.assertEqual(etl.find_last_value(ref, 'Y', 'c'), 'Y7')

    def test_find_in_header(self):
        self.assertEqual(etl.find_in_header(ref, 'TESTINGCOL'), 'J')       

    def test_colvals_notincol(self):
        self.assertEqual(tuple(etl.colvals_notincol(db, 'A', ref, 'A')), 
            tuple(['notincluded1','notincluded2','notincluded3']))

    def test_report_a_log(self):
        etl.report_a_log('**FILE 1**', 'file1')
        etl.report_a_log('ent2', 'file1')
        etl.report_a_log('ent2', 'file1')
        etl.report_a_log('**FILE 2**', 'file2')
        etl.report_a_log('ent5', 'file2')
        etl.report_a_log('ent5', 'file2')
        etl.report_a_log('**FILE 3**', 'file3')
        etl.report_a_log('ent5', 'file3')
        etl.report_a_log('ent5', 'file3')

        etl.report_a_log('', 'text')

        assert os.path.exists('/Users/ewanog/code/nepal-earthquake/shelter/etl/etl/logs/cleaned_log.txt')

    def test_consolidate_specific(self):
        #scenarios are: new agency, agency that is > 80pct inserted, agency <80 pct
        #create historical db
        db = Workbook().active
        db.append(("Implementing agency", "dummy", "Additional Comments", "Last Update"))

        #other agency
        for i in xrange(10):
            db.append(('agency_not_inserting','dummy','add','6-16-90'))

        #agency over 80
        for i in xrange(5):
            db.append(('agency_existing_over_80','dummy','add','6-16-90'))

        #agency under 80
        for i in xrange(40):
            db.append(('agency_existing_under_80','dummy','add','6-16-90'))

        for i in xrange(40):
            db.append(('Government','dummy','add','6-16-90'))


        #create agency not in db
        wb1 = Workbook()
        wb1.create_sheet(2, 'Distributions')
        wb1.create_sheet(3, 'Trainings')
        ws1 = wb1.get_sheet_by_name('Distributions')
        ws1.append(("Implementing agency", "dummy", "Additional Comments"))
        wb1.get_sheet_by_name('Trainings').append(("Implementing agency", "dummy", "Additional Comments"))
        wb1.get_sheet_by_name('Trainings').append(("agency", "dummy", "etc"))
        for i in xrange(5):
            ws1.append(("madnewagency", "dummy"))

        #create agency in db >80
        wb2 = Workbook()
        wb2.create_sheet(2, 'Distributions')
        wb2.create_sheet(3, 'Trainings')
        ws2 = wb2.get_sheet_by_name('Distributions')
        ws2 .append(("Implementing agency", "dummy", "Additional Comments"))
        wb2.get_sheet_by_name('Trainings').append(("Implementing agency", "dummy", "Additional Comments"))
        wb2.get_sheet_by_name('Trainings').append(("agency", "dummy", "etc"))
        for i in xrange(50):
            ws2.append(("agency_existing_over_80", "dummy"))

        #create agency in db <80
        wb3 = Workbook()
        wb3.create_sheet(2, 'Distributions')
        wb3.create_sheet(3, 'Trainings')
        ws3 = wb3.get_sheet_by_name('Distributions')
        ws3.append(("Implementing agency", "dummy", "Additional Comments"))
        wb3.get_sheet_by_name('Trainings').append(("Implementing agency", "dummy", "Additional Comments"))
        wb3.get_sheet_by_name('Trainings').append(("agency", "dummy", "etc"))
        for i in xrange(4):
            ws3.append(("agency_existing_under_80", "dummy"))

        #create another agency NOT in db
        wb4 = Workbook()
        wb4.create_sheet(2, 'Distributions')
        wb4.create_sheet(3, 'Trainings')
        ws4 = wb4.get_sheet_by_name('Distributions')
        ws4 .append(("Implementing agency", "dummy", "Additional Comments"))
        wb4.get_sheet_by_name('Trainings').append(("Implementing agency", "dummy", "Additional Comments"))
        wb4.get_sheet_by_name('Trainings').append(("agency", "dummy", "etc"))
        for i in xrange(25):
            ws4.append(("datnewnew", "dummy"))

        #gov
        wb5 = Workbook()
        wb5.create_sheet(2, 'Distributions')
        wb5.create_sheet(3, 'Trainings')
        ws5 = wb5.get_sheet_by_name('Distributions')
        ws5 .append(("Implementing agency", "dummy", "Additional Comments"))
        wb5.get_sheet_by_name('Trainings').append(("Implementing agency", "dummy", "Additional Comments"))
        wb5.get_sheet_by_name('Trainings').append(("agency", "dummy", "etc"))
        for i in xrange(3):
            ws5.append(("Government", "dummy"))



        #final counts should be:
        #other agency: 10 (agency_not_inserting)
        #agency over 80: 50 (agency_existing_over_80)
        #agnecy under 80: 40 (agency_existing_under_80)
        #new agency: 25 (datnewnew)
        #other new agency: 5 (madnewagency)

        cons = etl.consolidate(db, ((wb1,'f'), (wb2,'f'), (wb3,'f'), (wb4,'f'), (wb5,'f')))
        cons_sheet = cons.get_sheet_by_name('Distributions')
        etl.print_sheet(cons_sheet)
        r = etl.get_values(cons_sheet.columns[0])
        c = Counter(r)
        print c
        print etl.get_values(cons_sheet.rows[1])

        self.assertEqual(c['agency_not_inserting'], 10)
        self.assertEqual(c['agency_existing_over_80'], 50)
        self.assertEqual(c['agency_existing_under_80'], 40)
        self.assertEqual(c['datnewnew'], 25)
        self.assertEqual(c['Government'], 43)
        self.assertEqual(len(r), 173+1) #+1 for header

    def test_mismatch_header_cons(self):
       #create agency not in db
        wb1 = Workbook()
        wb1.create_sheet(2, 'Distributions')
        ws1 = wb1.get_sheet_by_name('Distributions')
        ws1 .append(("not ", "matching"))
        for i in xrange(5):
            ws1.append(("madnewagency", "dummy"))

        #create agency in db >80
        wb2 = Workbook()
        wb2.create_sheet(2, 'Distributions')
        ws2 = wb1.get_sheet_by_name('Distributions')
        ws2 .append(("no", "way"))
        for i in xrange(50):
            ws2.append(("agency_existing_over_80", "dummy"))

        r = etl.consolidate(wb1,(wb2))
        self.assertEqual(len(r.rows), 5)

    def test_get_values(self):
        db = Workbook().active
        db.append(('val', 'blank', 'key'))
        db.append(('val', 'key'))
        self.assertEqual(etl.get_values(db.rows[1]),['val','key'])

    def test_keep_dict(self):
        """row is new value, dict is old"""
        db = Workbook().active
        db.append(('Activity Status', 'Completion Date\n (Actual or Planned)', 'Start date \n(Actual or Planned)'))
        db.append(('Completed', '12/15/2015','12/10/2015'))
        d_v =  ('Completed', '12/15/2015','12/12/2015')
        self.assertTrue(etl.keep_dict(db.rows[1], d_v, db))

    def test_none_row_true(self):
        test = Workbook().active
        test.append(('None','None','None','None','None','None'))
        self.assertTrue(etl.none_row(test.rows[0]))

    def test_none_row_false(self):
        test = Workbook().active
        test.append(('Not','repeating','!!',';)',':D'))
        self.assertFalse(etl.none_row(test.rows[0]))

    def test_split_get_sheets(self):
        test = Workbook().active
        test.append(("Implementing agency", "dummy", "Additional Comments"))
        for i in xrange(40):
            test.append(("ag1", "dummy", "dummy"))
        for i in xrange(20):
            test.append(("ag2", "dummy", "dummy"))
        for i in xrange(10):
            test.append(("ag1", "dummy", "dummy"))
        for i in xrange(80):
            test.append(("ag3", "dummy", "dummy"))

        rs = etl.split_get_sheets(test)
        self.assertEqual(rs[0][0], 'ag1')
        self.assertEqual(len(rs[0][1]), 50)
        self.assertEqual(rs[1][0], 'ag2')
        self.assertEqual(len(rs[1][1]), 20)
        self.assertEqual(rs[2][0], 'ag3')
        self.assertEqual(len(rs[2][1]), 80)

    def test_clean_output(self):
        test = Workbook().active
        test.append(("# Items / # Man-hours / NPR", "Completion Date", "not_either"))
        test.append(('1','6/16/90','etc'))
        res = etl.clean_output(test)
        self.assertTrue(isinstance(res.rows[1][0].value,int))
        self.assertTrue(res.rows[1][1].value == '16/06/90')

    def test_copy_sheet(self):
        to = Workbook().active
        to.append(("ag1", "dummy", "dummy"))
        fro = Workbook().active
        fro.append(("ag1", "dummy", "dummy"))
        etl.copy_sheet(to,fro)

    def test_get_unsure_ws(self):
        wb = Workbook()
        wb.create_sheet(2, 'Distributions')
        wb.create_sheet(3, 'Trainings')
        wb.create_sheet(4, 'thisone')
        self.assertEqual(etl.get_unsure_ws(wb, ['d','bb','thisone']).title,'thisone')

    def test_consolidate_specfic(self):
        #INCOMPLETE!!!
        ws = etl.pull_wb('/Users/ewanog/tmp/cleaned//C - Navajyoti Center.xlsx', 'local', True)
        r = etl.consolidate_specfic(etl.pull_wb('/Users/ewanog/tmp/simp.xlsx', 'local', True), [(ws,'f')], 'Trainings')

    def test_add_vals_row(self):
        ws = Workbook().active
        ws.append(("head1", "head2", "head3"))
        ws.append(("vals", "morevals", "allthevals"))
        r = etl.add_vals(ws, ['new','header'], 'D1','r')
        rh = etl.get_values(r.rows[0])
        self.assertEqual(rh, ['head1', 'head2', 'head3', 'new','header'])

    def test_add_vals_col(self):
        ws = Workbook().active
        r  = etl.add_vals(ws, ['new','vals'], 'B1','c')
        rh = etl.get_values(r.columns[1])
        self.assertEqual(rh, ['new','vals'])

    def test_add_meta_colors(self):
        ws = Workbook().active
        ws.append(('etc','Additional Comments','not','blank'))

        template = Workbook().active
        template['A1'].fill = PatternFill(start_color = 'FFFFFFFF')
        ws = etl.add_meta_colors(ws, template)
        for v in ws.rows[0]:
            print v.value
            print v.style
            print
            print
            print

if __name__ == '__main__':
    unittest.main()
