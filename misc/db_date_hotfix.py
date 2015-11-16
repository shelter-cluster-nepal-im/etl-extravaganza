__author__ = 'ewanog'

import os
import shutil
import etl
from openpyxl import Workbook

files = ['/Users/ewanog/Downloads/tm/'+f for f in \
         os.listdir('/Users/ewanog/Downloads/tm/') if '.xlsx' in f and '~' not in f and 'merge.xlsx' not in f]

print files

d = []
t = []

it = 0
for f in files:
    w = etl.pull_wb(f, 'local',True)
    d.append(etl.get_all_values_from_ws(w.get_sheet_by_name('Distributions'), 'r'))
    t.append(etl.get_all_values_from_ws(w.get_sheet_by_name('Training'), 'r'))

w = Workbook()
w.create_sheet(1,'Distributions')
w.create_sheet(2,'Training')


sum = 0
print 'Dist'
s = w.get_sheet_by_name('Distributions')
s.append(d[0][0])
for entry in d:
    print len(entry) - 1
    sum += len(entry) - 1
    for r in entry[1:]:
        s.append(r)

print 'sum dist: ' + str(sum)

sum = 0
print 'Train'
s = w.get_sheet_by_name('Training')
s.append(t[0][0])
for entry in t:
    print len(entry) - 1
    sum += len(entry) - 1
    for r in entry[1:]:
        s.append(r)

print 'sum train: ' + str(sum)


etl.send_wb('/Users/ewanog/Downloads/tm/merge.xlsx', w, 'local')