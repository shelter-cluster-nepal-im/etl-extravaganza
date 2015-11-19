import unittest
import etl
from exql import upload_a_wb
from sqlalchemy import Column, Integer, String, Float
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import relationship
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import column_index_from_string
import exql

Base = declarative_base()
engine = create_engine(os.environ['dbk'])
Base.metadata.bind = engine
DBSession = sessionmaker(bind=engine)
session = DBSession()
Base.metadata.create_all(engine)


class TestUpload(unittest.TestCase):

    def test_append(self):
        wb = Workbook()
        ret = xls_to_sql.get_sheets([wb], ('sheet1', 'sheet3'))
        self.assertEqual([v.title for v in xls_to_sql.get_sheets([wb], ('sheet1', 'sheet3'))] \
                         , ['sheet1','sheet3'])

    def test_create_with_sheet(self):
        s = Workbook().active
        sht = Sheet(sht_raw, sheet_name)


    def test_get_sht_nm(self):
        w = Workbook().active
        w.get_active_sheet().title = 'new'

        self.assertEqual(upload_a_wb.get_sht_nm(w, None), 'new'))


class test(Base):
    __tablename__ = 'test'
    targeting = Column(String(250), primary_key=True)
    quantity = Column(Integer)
    total_hh = Column(Integer)
    float = Column(Float)

if __name__ == '__main__':
    unittest.main()
