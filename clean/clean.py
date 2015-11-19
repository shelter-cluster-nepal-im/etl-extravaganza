"""collection of cleaning and verification algos"""

#notes: two admin1_dist, using 'E'
#TODO: create a versatile stystem of creating filters that automitaiclly creates tests
#Nones?
#return col and cols methods

from etl import etl
import string
from datetime import datetime
from dateutil.parser import *
from openpyxl.cell import column_index_from_string
from datetime import datetime

def return_message(message, list):
    """return message and list"""
    return [message, list]

def algo1(db, ref):
    #***Column A must be in Reference>ImplementingAgency if not: make a note
    db_col_loc = etl.find_in_header(db, 'Implementing agency')
    ref_col_loc = etl.find_in_header(ref,'Implementing_Agency_Name')
    missing_names = etl.colvals_notincol(db, db_col_loc, ref, ref_col_loc)

    return db, ref, return_message('Agencies not in the reference:' ,missing_names)

def algo2(db,ref):
    #Column B: If == Implementing Agency set Column B=Internal
    #assumes they are next to each other
    db_col_loc_impl = etl.find_in_header(db, 'Implementing agency')
    db_col_loc_source = etl.find_in_header(db, 'Sourcing Agency')
    vals_changed = []

    for row in db.iter_rows(db_col_loc_impl + "2:" +
        etl.find_last_value(db, db_col_loc_source, 'c')):
            if row[0].value == row[1].value:
                vals_changed.append(etl.xstr(row[1].value))
                row[1].value = 'INTERNAL'

    return db, ref, return_message('Sourcing Agencies set as INTERNAL:' ,vals_changed)


def algo3(db,ref):
    #* what return? change?
    #Column C: must be in Reference>LocalPartnerAgency
    #If not: Check if mispelling (like '%agencyname%)

    db_col_loc = etl.find_in_header(db, 'Local partner agency')
    ref_col_loc = etl.find_in_header(ref,'Local_Partner_Agency')
    missing_names = etl.colvals_notincol(db, db_col_loc, ref, ref_col_loc)

#    return db, ref, return_message('' ,)

def algo4(db,ref):
    #if Column C == Column A
    #Clear column C

    db_col_loc_a = etl.find_in_header(db, 'Implementing agency')
    db_col_loc_c = etl.find_in_header(db, 'Local partner agency')
    vals_changed = []

    for row in db.iter_rows(db_col_loc_a + "2:" +
        etl.find_last_value(db, db_col_loc_c, 'c')):
            if etl.xstr(row[0].value) == etl.xstr(row[2].value):
                vals_changed.append(etl.xstr(row[2].value))
                row[2].value = ''

    return db, ref, return_message('Local Partner Agencies matched Implementing and were cleared' ,vals_changed)

def algo5(db,ref):
    #if Column C contains Column A: remove all spaces, commas, dashes and the matching substring in C
    db_col_loc_a = etl.find_in_header(db, 'Implementing agency')
    db_col_loc_c = etl.find_in_header(db, 'Local partner agency')
    vals_changed = []

    for row in db.iter_rows(db_col_loc_a + "2:" +
        etl.find_last_value(db, db_col_loc_c, 'c')):
            if etl.xstr(row[0].value) in etl.xstr(row[2].value) and etl.xstr(row[0].value) != etl.xstr(row[2].value):
                vals_changed.append(row[2].value)
                row[2].value = row[2].value.replace(row[0].value,'').replace(' ','').replace(',','').replace('-','')

    return db, ref, return_message('IA contained in LPA for the following:' ,vals_changed)

def algo6(db,ref):
    #Column G: must be in Reference>District
    db_col_loc = etl.find_in_header(db, 'District')
    ref_col_loc = etl.find_in_header(ref,'Admin1_District')
    missing_names = etl.colvals_notincol(db, db_col_loc, ref, ref_col_loc)
    return db, ref, return_message('Districts not in the reference:' ,missing_names)

def algo7(db,ref):
    #Column H: must be in Reference>VDC - relative to districy
    #the district in which a vdc in the db is located must = the district in ref
    db_dist_loc = etl.find_in_header(db, 'District')
    db_vdc_loc = etl.find_in_header(db, 'VDC / Municipalities')
    ref_dist_loc = etl.find_in_header(ref, 'Admin1_District')
    ref_vdc_loc = etl.find_in_header(ref, 'Admin2_OCHA_VDC-Municipality')

    #create tuple of vdc, dist for db and ref
    db_tuple = []
    for row in db.iter_rows(db_dist_loc + "2:" +
        etl.find_last_value(db, db_vdc_loc, 'c')):
            db_tuple.append((etl.xstr(row[0].value), etl.xstr(row[1].value)))

    ref_tuple = []
    #do it just for dist, then for vdc as they're not neighbors
    for row in ref.iter_rows('E2:' +
        etl.find_last_value(ref, 'E', 'c')):
            ref_tuple.append((etl.xstr(row[0].value),))




    it = 0
    for row in ref.iter_rows(ref_vdc_loc + "2:" +
        etl.find_last_value(ref, ref_vdc_loc, 'c')):
            ref_tuple[it]+=(etl.xstr(row[0].value.encode("utf8")),)
            it+=1

    #iterate through db_tuple and see if tuple is in ref
    malformed_vdc = []
    for tup in db_tuple:
        if tup not in ref_tuple:
            malformed_vdc.append(etl.xstr(tup))

    return db, ref, return_message('VDCs in incorrect districts:' ,malformed_vdc)

def algo8(db,ref):
    #Column I: must be in reference>Ward or a number or numbers separated by commas

    #take all into missing names and trim from there
    db_col_loc = etl.find_in_header(db, 'Municipal Ward')
    ref_col_loc = etl.find_in_header(ref,'Wards')
    missing_names = etl.colvals_notincol(db, db_col_loc, ref, ref_col_loc)

    #trim missing_names to see if they contain any forbidden chars
    invalid = []

    for v in missing_names:
        for letter in v:
            #if we've found an illegal letter
            if letter not in set(string.digits + ' ' + ','):
                invalid.append('(' + v + ')')
                break


    return db, ref, return_message('Malformed wards: ' ,invalid)

def algo9(db,ref):
    #column J: must be in reference>Type of Activity
    db_col_loc = etl.find_in_header(db, 'Action type')
    ref_col_loc = etl.find_in_header(ref,'Action_Type')
    missing_names = etl.colvals_notincol(db, db_col_loc, ref, ref_col_loc)

    return db, ref, return_message('Incorrect Action types:' ,missing_names)

def algo10(db,ref):
    #column J: must be in reference>ActionDescription relative to Action_Type
    #There are 4 action types each with their own self labeled column
    db_type_loc = etl.find_in_header(db, 'Action type')
    db_desc_loc = etl.find_in_header(db, 'Action description')
    ref_col_loc = etl.find_in_header(ref,'Action_Type')

    #get ref action type tuples
    act_types_tups = []

    #go through action types and go through respected cols and make tuples
    for row in ref.iter_rows(ref_col_loc + "2:" +
        etl.find_last_value(ref, ref_col_loc, 'c')):
            act = etl.xstr(row[0].value)
            act_col = etl.find_in_header(ref, act)
            for row in ref.iter_rows(act_col + "2:" +
                etl.find_last_value(ref, act_col, 'c')):
                    act_types_tups.append((act, etl.xstr(row[0].value)))

    #get tuples of dist actions
    type_desc_tup = []
    for row in db.iter_rows(db_type_loc + "2:" +
        etl.find_last_value(db, db_desc_loc, 'c')):
            type_desc_tup.append((etl.xstr(row[0].value),etl.xstr(row[1].value)))

    malformed_acts = []
    #go through dist actions and see if they're in ref tuples
    for tup in type_desc_tup:
        if tup not in act_types_tups:
            malformed_acts.append(etl.xstr(tup))

    return db, ref, return_message('Action descriptions not in action types:' ,malformed_acts)

def algo11(db,ref):
    #Column L: must be in Reference>Target
    db_col_loc = etl.find_in_header(db, 'Targeting')
    ref_col_loc = etl.find_in_header(ref,'Target')
    missing_names = etl.colvals_notincol(db, db_col_loc, ref, ref_col_loc)

    return db, ref, return_message('Targets not in Referece' ,missing_names)

def algo12(db,ref):
    #Column M: must be a number>=0 and (more conditions to follow)
    db_col_loc = etl.find_in_header(db, '# Items / # Man-hours / NPR')
    bad_vals = []

    for row in db.iter_rows(db_col_loc + "2:" +
        etl.find_last_value(db, db_col_loc, 'c')):
            if not etl.xstr(row[0].value).isdigit():
                if row[0].value == None:
                    bad_vals.append('Blank @ cell' + row[0].coordinate)
                else:
                    bad_vals.append(etl.xstr(row[0].value))

    return db, ref, return_message('Invalid numbers for # Items / # Man-hours / NPR' ,bad_vals)

def algo13(db,ref):
    #Column N: must be a number>=0 and (more conditions to follow)
    db_col_loc = etl.find_in_header(db, 'Total Number Households')
    bad_vals = []

    for row in db.iter_rows(db_col_loc + "2:" +
        etl.find_last_value(db, db_col_loc, 'c')):
            if not etl.xstr(row[0].value).isdigit():
                if row[0].value == None:
                    bad_vals.append('Blank@ ' + row[0].coordinate)
                else:
                    bad_vals.append(etl.xstr(row[0].value))

    return db, ref, return_message('Invalid numbers for Total Number Households' ,bad_vals)

def algo14(db,ref):
    #Column O: must be a number>=0 and (more conditions to follow)
    db_col_loc = etl.find_in_header(db, 'Average cost per households (NPR)')
    bad_vals = []

    for row in db.iter_rows(db_col_loc + "2:" +
        etl.find_last_value(db, db_col_loc, 'c')):
            if not etl.xstr(row[0].value).isdigit():
                if row[0].value == None:
                    bad_vals.append('Blank@ ' + row[0].coordinate)
                else:
                    bad_vals.append(etl.xstr(row[0].value))


    return db, ref, return_message('Invalid numbers for Average cost per households (NPR)' ,bad_vals)

def algo15(db,ref):
    #Column P: must be a number>=0 and <= Column N (more conditions to follow)
    db_cnt_loc = etl.find_in_header(db, 'Total Number Households')
    db_fem_loc = etl.find_in_header(db, 'Female headed households')


    #get count values
    cnt_vals = []
    for row in db.iter_rows(db_cnt_loc + "2:" +
        etl.find_last_value(db, db_cnt_loc, 'c')):
            cnt_vals.append(etl.xstr(row[0].value))

    #get female values
    fem_vals = []
    for row in db.iter_rows(db_fem_loc + "2:" +
        etl.find_last_value(db, db_fem_loc, 'c')):
            fem_vals.append(etl.xstr(row[0].value))

    malformatted = []
    #zip and check values
    for vals in zip(cnt_vals, fem_vals):
        if not (vals[0].isdigit() and vals[1].isdigit()):
            malformatted.append(etl.xstr(vals))
        elif int(vals[1]) > int(vals[0]):
            malformatted.append(etl.xstr(vals))

    return db, ref, return_message('Total #HH and Female #HH conflict or issue ' ,malformatted)

def algo16(db,ref):
    #Column Q: must be a number>=0 and <= Column N (more conditions to follow)
    db_cnt_loc = etl.find_in_header(db, 'Total Number Households')
    db_vul_loc = etl.find_in_header(db, 'Vulnerable Caste / Ethnicity households ')


    #get count values
    cnt_vals = []
    for row in db.iter_rows(db_cnt_loc + "2:" +
        etl.find_last_value(db, db_cnt_loc, 'c')):
            cnt_vals.append(etl.xstr(row[0].value))

    #get female values
    vul_vals = []
    for row in db.iter_rows(db_vul_loc + "2:" +
        etl.find_last_value(db, db_vul_loc, 'c')):
            vul_vals.append(etl.xstr(row[0].value))

    malformatted = []
    #zip and check values
    for vals in zip(cnt_vals, vul_vals):
        if not (vals[0].isdigit() and vals[1].isdigit()):
            malformatted.append(etl.xstr(vals))
        elif int(vals[1]) > int(vals[0]):
            malformatted.append(etl.xstr(vals))

    return db, ref, return_message('Total #HH and Vulnerable #HH conflict or issue ' ,malformatted)

def algo17(db,ref):
    #column R: must be in reference>Status
    db_col_loc = etl.find_in_header(db, 'Activity Status')
    ref_col_loc = etl.find_in_header(ref,'Status')
    missing_names = etl.colvals_notincol(db, db_col_loc, ref, ref_col_loc)

    return db, ref, return_message('Activity Status not in Reference' ,missing_names)

def algo18(db,ref):
    #Column S: must be a date>= 25/04/2015
    date_col_loc = column_index_from_string(etl.find_in_header(db, 'DD - Start')) - 1

    cur = ""
    bad_date = []

    for row in db.rows[1:]:
        try:
            cur = datetime(day = int(etl.xstr(row[date_col_loc].value)), \
                            month = int(etl.xstr(row[date_col_loc+1].value)), \
                            year = int(etl.xstr(row[date_col_loc+2].value)))

            if cur < parse('4-25-2015'):
                bad_date.append(str(cur) + ' @ row ' + str(row[0].row))

        except :
            bad_date.append(str(cur) + ' @ row ' + str( quit()[0].row))

    return db, ref, return_message('Malformatted date before EQ date', bad_date)



def algo19(db,ref):
    #Column T: must be a date>= 25/04/2015 and >=Column S
    start_col_loc = column_index_from_string(etl.find_in_header(db, 'DD - Start')) - 1
    comp_col_loc = column_index_from_string(etl.find_in_header(db, 'DD - End')) - 1

    bad_date = []
    start = ""
    comp = ""

    for row in db.rows[1:]:
        try:
            start = datetime(day = int(etl.xstr(row[start_col_loc].value)), \
                            month = int(etl.xstr(row[start_col_loc+1].value)), \
                            year = int(etl.xstr(row[start_col_loc+2].value)))

            comp = datetime(day = int(etl.xstr(row[comp_col_loc].value)), \
                            month = int(etl.xstr(row[comp_col_loc+1].value)), \
                            year = int(etl.xstr(row[comp_col_loc+2].value)))

            if comp < start:
                bad_date.append(str(start) + ' @ row ' + str(row[0].row))

        except:
            bad_date.append(str(start) + ' @ row ' + str(row[0].row))

    return db, ref, return_message('Malformatted date or before Start Date', bad_date)

def algo20(db,ref):
    #Output if completion date < current date and activity status not like completed
    status_col_loc = column_index_from_string(etl.find_in_header(db, 'Activity Status')) - 1
    comp_col_loc = column_index_from_string(etl.find_in_header(db, 'DD - End')) - 1

    bad_date = []
    valid = True
    comp = ""

    for row in db.rows[1:]:
        try:
            comp = datetime(day = int(etl.xstr(row[comp_col_loc].value)), \
                month = int(etl.xstr(row[comp_col_loc+1].value)), \
                year = int(etl.xstr(row[comp_col_loc+2].value)))
        except:
            valid = False

        if valid:
            try:
                if comp < datetime.now() and 'completed' not in row[status_col_loc].value.lower():
                    bad_date.append('%s row %i' % (str(comp),row[0].row))
            except:
                    bad_date.append('%s row %i' % (str(comp),row[0].row))

    return db, ref, return_message('Bad date or Completion Date has passed for ' ,bad_date)

def algo21(db,ref):
    #
    return db, ref, return_message('' ,)

def algo22(db,ref):
    #
    return db, ref, return_message('' ,)

def algo23(db,ref):
    #
    return db, ref, return_message('' ,)

def algo24(db,ref):
    #
    return db, ref, return_message('' ,)

def algo25(db,ref):
    #
    return db, ref, return_message('' ,)

def algo26(db,ref):
    #
    return db, ref, return_message('' ,)

def algo27(db,ref):
    #
    return db, ref, return_message('' ,)

def algo28(db,ref):
    #
    return db, ref, return_message('' ,)

def algo29(db,ref):
    #
    return db, ref, return_message('' ,)

def algo30(db,ref):
    #
    return db, ref, return_message('' ,)

def algo31(db,ref):
    #
    return db, ref, return_message('' ,)

def algo32(db,ref):
    #
    return db, ref, return_message('' ,)

def algo33(db,ref):
    #
    return db, ref, return_message('' ,)

def algo34(db,ref):
    #
    return db, ref, return_message('' ,)

def algo35(db,ref):
    #
    return db, ref, return_message('' ,)

def algo36(db,ref):
    #
    return db, ref, return_message('' ,)

def algo37(db,ref):
    #
    return db, ref, return_message('' ,)

def algo38(db,ref):
    #
    return db, ref, return_message('' ,)

def algo39(db,ref):
    #
    return db, ref, return_message('' ,)

def algo40(db,ref):
    #
    return db, ref, return_message('' ,)

def algo41(db,ref):
    #
    return db, ref, return_message('' ,)

def algo42(db,ref):
    #
    return db, ref, return_message('' ,)

def algo43(db,ref):
    #
    return db, ref, return_message('' ,)

def algo44(db,ref):
    #
    return db, ref, return_message('' ,)

def algo45(db,ref):
    #
    return db, ref, return_message('' ,)

def algo46(db,ref):
    #
    return db, ref, return_message('' ,)

def algo47(db,ref):
    #
    return db, ref, return_message('' ,)

def algo48(db,ref):
    #
    return db, ref, return_message('' ,)



def algo(db,ref):
    print 'x'